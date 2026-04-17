from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
OUTPUT_PATH = ROOT / "docs" / "ops" / "CEO_WEEKLY_SCORECARD.xlsx"


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E7F5")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF4FB")
BOLD = Font(bold=True)


def set_widths(ws: Workbook.worksheets, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_row(cell_range) -> None:
    for cell in cell_range:
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_overview_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Overview"
    rows = [
        ("Document", "CEO Weekly Scorecard"),
        ("Purpose", "Invulbare werklaag voor wekelijkse CEO-sturing zonder Markdown als invoervlak."),
        ("How to use", "Werk deze scorecard wekelijks bij. Gebruik alleen echte cijfers of noteer 'nog niet gemeten'."),
        ("Dominant phase", "Phase 1 - Revenue Proof Engine"),
        ("Parallel phase", "Phase 2 - Delivery Stability Engine"),
        ("Core rule", "Focus op revenue + proof en delivery stability; geen nieuwe grote projecten buiten deze command window."),
        ("Lead source of truth", "Verisight_CRM.xlsx"),
        ("Delivery source of truth", "/beheer en /campaigns/[id]"),
        ("Mirror rule", "Deals en Clients zijn weekmirrors; werk eerst het primaire systeem bij."),
    ]
    for row in rows:
        ws.append(row)
    style_row(ws[1])
    set_widths(ws, {"A": 22, "B": 110})


def build_weekly_focus_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Weekly Focus")
    ws.append(["Field", "Value", "Prompt"])
    style_row(ws[1])

    rows = [
        ("Week van", "", "Kies alleen de week die nu loopt."),
        ("Deze week telt vooral", "", "Welke ene beweging maakt deze week succesvol?"),
        ("Belangrijkste commerciële push", "", "Welk gesprek, voorstel of outbound-batch moet deze week echt bewegen?"),
        ("Belangrijkste delivery-risk", "", "Welk klanttraject vraagt nu de meeste aandacht?"),
        ("Topprioriteit 1", "", "Hou het concreet en uitvoerbaar."),
        ("Topprioriteit 2", "", "Alleen invullen als je het deze week echt doet."),
        ("Topprioriteit 3", "", "Alleen invullen als je het deze week echt doet."),
        ("Niet doen 1", "", "Wat laat je deze week bewust liggen?"),
        ("Niet doen 2", "", "Wat laat je deze week bewust liggen?"),
        ("Niet doen 3", "", "Wat laat je deze week bewust liggen?"),
        ("Besluit deze week", "", "Welk besluit moet expliciet in de decision log landen?"),
    ]
    for row in rows:
        ws.append(row)

    set_widths(ws, {"A": 28, "B": 38, "C": 68})
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_scorecard_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Weekly Scorecard")
    ws.append(["Section", "Metric", "Value", "Notes"])
    style_row(ws[1])

    rows = [
        ("Current frame", "Week van", "", ""),
        ("Current frame", "Dominante fase", "Phase 1 - Revenue Proof Engine", ""),
        ("Current frame", "Parallelle fase", "Phase 2 - Delivery Stability Engine", ""),
        ("Current frame", "Owner", "Founder-led", ""),
        ("Current frame", "Lead source of truth", "Verisight_CRM.xlsx", "Werk eerst de CRM-pipeline bij en spiegel daarna naar Deals."),
        ("Current frame", "Delivery source of truth", "/beheer en /campaigns/[id]", "Werk eerst de app-deliverylaag bij en spiegel daarna naar Clients."),
        ("Revenue proof", "Gekwalificeerde gesprekken deze week", "", ""),
        ("Revenue proof", "Lopende proposals/offertes", "", ""),
        ("Revenue proof", "Betaalde klanten of pilots in flight", "", ""),
        ("Revenue proof", "Win rate rolling 30 dagen", "", ""),
        ("Revenue proof", "Doorlooptijd gesprek naar voorstel", "", ""),
        ("Revenue proof", "Nieuwe case-proof snippets deze week", "", ""),
        ("Revenue proof", "Grootste bezwaar", "", ""),
        ("Revenue proof", "ICP-kwaliteit", "", "sterk / gemengd / diffuus"),
        ("Delivery stability", "Actieve klanttrajecten", "", ""),
        ("Delivery stability", "Trajecten met hoog risico", "", ""),
        ("Delivery stability", "Delivery defects deze week", "", ""),
        ("Delivery stability", "Handmatige escalaties", "", ""),
        ("Delivery stability", "Dagen akkoord naar live start", "", ""),
        ("Delivery stability", "Dagen live start naar eerste managementwaarde", "", ""),
        ("Delivery stability", "Trajecten met expliciete vervolgstap", "", ""),
        ("Cash and capacity", "Verwachte omzet komende 30 dagen", "", ""),
        ("Cash and capacity", "Verwachte omzet komende 90 dagen", "", ""),
        ("Cash and capacity", "Openstaande facturatie of cashrisico", "", ""),
        ("Cash and capacity", "Founder-capaciteit komende 14 dagen", "", ""),
        ("Cash and capacity", "Initiatieven buiten kernroute", "", ""),
    ]
    for row in rows:
        ws.append(row)

    set_widths(ws, {"A": 22, "B": 42, "C": 24, "D": 58})
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def build_deals_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Deals")
    ws.append(["Company", "Contact", "Route", "Stage", "Owner", "Next action", "Due date", "Biggest objection", "Notes"])
    style_row(ws[1])
    set_widths(ws, {"A": 24, "B": 24, "C": 18, "D": 18, "E": 16, "F": 32, "G": 16, "H": 30, "I": 42})
    ws.freeze_panes = "A2"


def build_clients_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Clients")
    ws.append(
        [
            "Client",
            "Route",
            "Stage",
            "Risk",
            "Owner",
            "Intake complete",
            "Import ready",
            "Activation ready",
            "Report delivered",
            "First management use",
            "Next step",
        ]
    )
    style_row(ws[1])
    set_widths(ws, {"A": 24, "B": 18, "C": 18, "D": 14, "E": 14, "F": 16, "G": 14, "H": 16, "I": 16, "J": 18, "K": 32})
    ws.freeze_panes = "A2"


def build_priorities_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Priorities")
    ws.append(["Type", "Item", "Why"])
    style_row(ws[1])
    entries = [
        ("Top priority", "", ""),
        ("Top priority", "", ""),
        ("Top priority", "", ""),
        ("Explicitly not doing", "", ""),
        ("Explicitly not doing", "", ""),
        ("Explicitly not doing", "", ""),
        ("Decision log follow-up", "", ""),
        ("Decision log follow-up", "", ""),
    ]
    for row in entries:
        ws.append(row)
    set_widths(ws, {"A": 22, "B": 42, "C": 62})
    ws.freeze_panes = "A2"


def build_operating_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Operating Rules")
    ws.append(["Type", "Primary system", "Weekly mirror", "Rule"])
    style_row(ws[1])
    entries = [
        ("Leads", "Verisight_CRM.xlsx", "Deals", "Deals is alleen een weekmirror van de live CRM-pipeline."),
        ("Delivery", "/beheer en /campaigns/[id]", "Clients", "Clients is alleen een weekmirror van de live deliverylaag."),
        ("Governance", "Weekly Scorecard", "", "Gebruik alleen echte data of noteer nog niet gemeten."),
    ]
    for row in entries:
        ws.append(row)
    set_widths(ws, {"A": 18, "B": 28, "C": 18, "D": 72})
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    build_overview_sheet(wb)
    build_weekly_focus_sheet(wb)
    build_scorecard_sheet(wb)
    build_deals_sheet(wb)
    build_clients_sheet(wb)
    build_priorities_sheet(wb)
    build_operating_rules_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
