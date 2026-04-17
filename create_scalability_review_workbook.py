from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
OUTPUT_PATH = ROOT / "docs" / "ops" / "SCALABILITY_REVIEW_WORKBOOK.xlsx"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E7F5")
BOLD = Font(bold=True)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header(row) -> None:
    for cell in row:
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_overview_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Overview"
    rows = [
        ("Document", "Scalability Review Workbook"),
        ("Purpose", "Werkboek voor de Verisight schaalbaarheidsreview en vervolgwaves."),
        ("Scope", "Hybrid: repo + Docs_External + operationele voorbeelden."),
        ("Horizon", "Komende 90 dagen binnen founder-led lean."),
        ("Current verdict", "Niet operationeel schaalbaar voorbij kleine founder-led throughput."),
        ("Core rule", "Eerst stabiliseren, dan systemiseren, pas daarna lichte automation."),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws[1])
    set_widths(ws, {"A": 22, "B": 110})


def build_domain_scores_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Domain Scores")
    ws.append(
        [
            "Domain",
            "Explicit route (0-2)",
            "Repeatable (0-2)",
            "Visible/measurable (0-2)",
            "Transferable (0-2)",
            "Scale pressure (0-2)",
            "Total",
            "Verdict",
            "Key evidence",
            "Biggest risk",
        ]
    )
    style_header(ws[1])
    rows = [
        ("Commercial scalability", 2, 1, 0, 1, 0, 4, "Rood", "Outbound playbook + prospectlijst bestaan; CRM-pipeline en scorecard deals zijn leeg.", "Pipeline draait nog te veel op geheugen en losse opvolging."),
        ("Delivery scalability", 2, 2, 1, 1, 0, 6, "Geel", "Onboarding playbook, procesflow en failure matrix zijn sterk; live queue- en checkpointdata ontbreken.", "Delegatie en 2x volume zijn nog niet bewezen."),
        ("Governance scalability", 2, 1, 0, 1, 0, 4, "Rood", "Cadans en templates bestaan; ingevulde scorecards en monthly reviews ontbreken.", "Bestuurlijk ritme blijft te kwetsbaar."),
        ("Source-of-truth scalability", 1, 1, 1, 1, 1, 5, "Geel", "Repo + register + workbook bestaan; leidende volgorde verschilt nog per document.", "Bronlagen kunnen onder druk weer uit elkaar lopen."),
        ("Operational tooling scalability", 1, 1, 0, 1, 1, 4, "Rood", "Sterke templates, maar weinig live metrics en geen centraal werkbord.", "Tooling ondersteunt invullen, niet sturen."),
        ("Founder dependency scalability", 1, 0, 0, 1, 1, 3, "Rood", "Founder blijft centrale triage- en herstelroute.", "Afwezigheid van 2 weken levert directe vertraging op."),
    ]
    for row in rows:
        ws.append(row)
    set_widths(ws, {"A": 28, "B": 16, "C": 16, "D": 18, "E": 16, "F": 16, "G": 10, "H": 12, "I": 54, "J": 54})
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_evidence_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Evidence & Cases")
    ws.append(["Source", "Type", "Observation", "Scalability signal", "Follow-up"])
    style_header(ws[1])
    rows = [
        ("Verisight_CRM.xlsx / Pipeline", "Ops evidence", "Alleen headers; geen actieve pipeline-items ingevuld.", "Negatief: commercieel ritme nog niet zichtbaar in systeem.", "Maak live pipeline verplicht werkritme."),
        ("Verisight_Prospectlijst_2026-04-10.xlsx", "Commercial seed list", "16 seed prospects met ICP-fit en volgende acties.", "Positief: ICP en outboundrichting zijn concreet.", "Verbind lijst direct met actieve pipeline."),
        ("CEO_WEEKLY_SCORECARD.xlsx / Deals", "Governance evidence", "Deals-tab leeg.", "Negatief: scorecard bestaat, maar ritme is nog niet bewezen.", "Wekelijks vullen en reviewen."),
        ("CEO_WEEKLY_SCORECARD.xlsx / Clients", "Delivery evidence", "Clients-tab leeg.", "Negatief: actieve trajecten zijn nog niet centraal zichtbaar.", "Voeg live checkpointoverzicht toe."),
        ("Verisight_Onboarding_Procesflow_2026-04-11.xlsx", "Delivery design", "Procesflow en checklist zijn helder uitgewerkt.", "Positief: deliveryroute is goed gedocumenteerd.", "Vertaal naar live operationsboard."),
        ("Decision Log", "Governance evidence", "3 besluiten vastgelegd.", "Gemengd: goede start, maar nog geen doorlopende ritmische besluitvorming.", "Na elke maandreview actualiseren."),
    ]
    for row in rows:
        ws.append(row)
    set_widths(ws, {"A": 34, "B": 18, "C": 44, "D": 44, "E": 34})
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_risks_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Top Risks")
    ws.append(["Risk", "Severity", "Why it matters now", "Mitigation wave"])
    style_header(ws[1])
    rows = [
        ("Lege pipeline-instrumentatie", "High", "Geen hard salesritme zichtbaar.", "Wave 1"),
        ("Founder als centrale triage-laag", "High", "Beperkt overdraagbaarheid en afwezigheidstolerantie.", "Wave 1"),
        ("Geen actieve throughputmeting", "High", "Schaalproblemen worden te laat gezien.", "Wave 1"),
        ("Governance vooral ontworpen", "High", "Cadans kan onder druk snel wegvallen.", "Wave 1"),
        ("Deliverycheckpoints niet live afgedwongen", "High", "Trajectrisico blijft te laat zichtbaar.", "Wave 1"),
        ("Source-of-truth nog niet hard genoeg", "Medium", "Groei verhoogt kans op dubbele waarheden.", "Wave 2"),
        ("Proof capture niet kernhard", "Medium", "Commercieel bewijs blijft kwetsbaar.", "Wave 2"),
        ("Geen duidelijke owner-laag naast founder", "Medium", "Delegatie blijft theoretisch.", "Wave 2"),
        ("Manual-first zonder WIP-limieten", "Medium", "Overbelasting wordt laat zichtbaar.", "Wave 2"),
        ("Automation te vroeg starten", "Medium", "Risico op verkeerde systeemcomplexiteit.", "Wave 3"),
    ]
    for row in rows:
        ws.append(row)
    set_widths(ws, {"A": 34, "B": 12, "C": 50, "D": 16})
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_waves_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Improvement Waves")
    ws.append(["Wave", "Goal", "Must deliver", "Explicitly not doing", "Review point"])
    style_header(ws[1])
    rows = [
        ("Wave 1 - Stabilize", "Governance- en operationslekken dichten.", "Live pipeline, actieve scorecard, maandreview, deliveryoverzicht, ingevulde capacity map.", "Geen automation, geen nieuw CRM, geen nieuwe productprojecten.", "Dag 30"),
        ("Wave 2 - Systemize", "Ownership, metrics en delegatie harder maken.", "Defectlog, rolling metrics, delegated run, proof checkpoints.", "Geen brede toolingbouw zonder gemeten pijn.", "Dag 60"),
        ("Wave 3 - Scale Readiness", "Alleen bewezen knelpunten standaardiseren.", "Automation shortlist met businesscase, duidelijke ops-/foundergrenzen.", "Geen Stripe-first of self-serve push.", "Dag 90"),
    ]
    for row in rows:
        ws.append(row)
    set_widths(ws, {"A": 24, "B": 36, "C": 52, "D": 44, "E": 14})
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    wb = Workbook()
    build_overview_sheet(wb)
    build_domain_scores_sheet(wb)
    build_evidence_sheet(wb)
    build_risks_sheet(wb)
    build_waves_sheet(wb)
    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
