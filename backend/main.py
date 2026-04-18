"""
Verisight — FastAPI Backend
===================================
Entrypoint: uvicorn backend.main:app --reload

Routes
------
  GET  /survey/{token}           → serve HTML survey form
  POST /survey/submit            → process & store submission
  GET  /api/org/{slug}/campaigns → list campaigns
  POST /api/org/{slug}/campaigns → create campaign
  POST /api/org/{slug}/campaigns/{id}/respondents → add respondents
  GET  /api/campaigns/{id}/stats → dashboard stats
  GET  /api/health               → health check
"""

from __future__ import annotations

import csv
import io
import logging
import os
from contextlib import asynccontextmanager

import sentry_sdk
from openpyxl import load_workbook
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration

_SENTRY_DSN = os.getenv("SENTRY_DSN")
if _SENTRY_DSN:
    sentry_sdk.init(
        dsn=_SENTRY_DSN,
        integrations=[FastApiIntegration(), SqlalchemyIntegration()],
        traces_sample_rate=0.2,   # 20% van requests getraceerd
        environment=os.getenv("ENVIRONMENT", "production"),
        # Zorg dat PII niet in Sentry belandt
        send_default_pii=False,
    )
from datetime import datetime, timezone
from pathlib import Path
from time import time
from typing import Annotated, Any
from uuid import UUID

from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy import text
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy.orm import Session, selectinload

from backend.database import DATABASE_URL, check_db_connection, get_db, init_db
from backend.email import send_contact_request_result, send_hr_notification, send_survey_invite
from backend.models import Campaign, ContactRequest, Organization, OrganizationSecret, Respondent, SurveyResponse
from backend.products.leadership.definition import DEFAULT_LEADERSHIP_MODULES
from backend.products.onboarding.definition import DEFAULT_ONBOARDING_MODULES
from backend.products.pulse.definition import DEFAULT_PULSE_MODULES
from backend.products.team.definition import DEFAULT_TEAM_MODULES
from backend.products.shared.registry import get_product_module
from backend.runtime import require_backend_admin_token, validate_runtime_config
from backend.scan_definitions import get_scan_definition
from backend.schemas import (
    CampaignCreate,
    CampaignRead,
    CampaignStats,
    ContactRequestCreate,
    ContactRequestRead,
    ContactRequestResponse,
    ContactRequestUpdate,
    InviteSendResult,
    OrganizationCreate,
    OrganizationRead,
    RespondentBatchCreate,
    RespondentCreate,
    RespondentImportIssue,
    RespondentImportPreviewRow,
    RespondentImportResponse,
    RespondentRead,
    SendInviteItem,
    SurveySubmit,
    SurveyResponseRead,
)
from backend.scoring import (
    anonymize_text,
    detect_patterns,
    ORG_FACTOR_KEYS,
)

# ---------------------------------------------------------------------------
# Optional OpenAI enrichment (only used for open text; graceful fallback)
# ---------------------------------------------------------------------------

# OpenAI-integratie tijdelijk uitgeschakeld (AVG — EU-werknemersdata mag niet
# zonder expliciete toestemming naar een Amerikaanse verwerker).
# Herinschakelen vereist: expliciete toestemming in surveyflow + DPA met OpenAI.
_openai_client = None
_openai_available = False


# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"

_IS_SQLITE = DATABASE_URL.startswith("sqlite")
_IS_PRODUCTION = os.getenv("ENVIRONMENT", "development").lower() == "production"
logger = logging.getLogger(__name__)


def _resolve_survey_modules(scan_type: str, enabled_modules: list[str] | None) -> list[str]:
    """Return only factor modules for the survey itself.

    Campaign-level add-ons, such as segment verdieping in the report, should
    never suppress factor questions in the survey. When no explicit factor
    subset is configured, the full default factor set remains active.
    """
    if not enabled_modules:
        if scan_type == "leadership":
            return list(DEFAULT_LEADERSHIP_MODULES)
        if scan_type == "onboarding":
            return list(DEFAULT_ONBOARDING_MODULES)
        if scan_type == "pulse":
            return list(DEFAULT_PULSE_MODULES)
        if scan_type == "team":
            return list(DEFAULT_TEAM_MODULES)
        return ORG_FACTOR_KEYS

    factor_modules = [module for module in enabled_modules if module in ORG_FACTOR_KEYS]
    if factor_modules:
        return factor_modules
    if scan_type == "leadership":
        return list(DEFAULT_LEADERSHIP_MODULES)
    if scan_type == "onboarding":
        return list(DEFAULT_ONBOARDING_MODULES)
    if scan_type == "pulse":
        return list(DEFAULT_PULSE_MODULES)
    if scan_type == "team":
        return list(DEFAULT_TEAM_MODULES)
    return ORG_FACTOR_KEYS


def _send_hr_notification_safe(
    *,
    to_email: str,
    campaign_name: str,
    campaign_id: str,
    total_completed: int,
    total_invited: int,
) -> None:
    try:
        send_hr_notification(
            to_email=to_email,
            campaign_name=campaign_name,
            campaign_id=campaign_id,
            total_completed=total_completed,
            total_invited=total_invited,
        )
    except Exception as exc:
        logger.warning("HR-notificatie mislukt: %s", exc)
        sentry_sdk.capture_message(
            f"HR-notificatie mislukt na survey-submit: {exc}",
            level="warning",
        )


@asynccontextmanager
async def lifespan(app: FastAPI):
    validate_runtime_config(is_production=_IS_PRODUCTION)
    # Alleen tabellen aanmaken bij SQLite (lokale dev zonder Supabase)
    # Bij PostgreSQL/Supabase beheert schema.sql de tabellen
    if _IS_SQLITE:
        init_db()
    yield


app = FastAPI(
    title="Verisight API",
    version="2.0.0",
    description="HR verloopanalyse platform — ExitScan & RetentieScan",
    lifespan=lifespan,
)

_FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:3000")
_ADDITIONAL_CORS_ORIGINS = [
    origin.strip()
    for origin in os.getenv("ADDITIONAL_CORS_ORIGINS", "").split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        _FRONTEND_URL,
        "http://localhost:3000",
        "http://localhost:3001",
        *_ADDITIONAL_CORS_ORIGINS,
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

EXIT_REASON_CODE_MAP = {
    "leiderschap": "P1",
    "cultuur": "P2",
    "groei": "P3",
    "beloning": "P4",
    "werkdruk": "P5",
    "rolonduidelijkheid": "P6",
    "beter_aanbod": "PL1",
    "carriere_switch": "PL2",
    "ondernemerschap": "PL3",
    "persoonlijk": "S1",
    "gezondheid": "S1",
    "verhuizing": "S2",
    "partner_verhuisd": "S2",
    "studie": "S3",
    "pensioen": "S3",
}

CONTACT_RATE_LIMIT_WINDOW_SECONDS = 15 * 60
CONTACT_RATE_LIMIT_MAX_REQUESTS = 5
_contact_request_buckets: dict[str, dict[str, float | int]] = {}

RESPONDENT_IMPORT_COLUMN_ALIASES = {
    "email": "email",
    "e-mail": "email",
    "e-mailadres": "email",
    "mailadres": "email",
    "department": "department",
    "afdeling": "department",
    "team": "department",
    "role_level": "role_level",
    "role level": "role_level",
    "functieniveau": "role_level",
    "niveau": "role_level",
    "functie_niveau": "role_level",
    "exit_month": "exit_month",
    "exit month": "exit_month",
    "exitmaand": "exit_month",
    "vertrekmaand": "exit_month",
    "uittreedmaand": "exit_month",
    "uittreedatum_maand": "exit_month",
    "uittreedatum maand": "exit_month",
    "annual_salary_eur": "annual_salary_eur",
    "salary": "annual_salary_eur",
    "jaarsalaris": "annual_salary_eur",
    "bruto_jaarsalaris": "annual_salary_eur",
    "bruto jaarsalaris": "annual_salary_eur",
}

ROLE_LEVEL_ALIASES = {
    "uitvoerend": "uitvoerend",
    "operationeel": "uitvoerend",
    "medewerker": "uitvoerend",
    "specialist": "specialist",
    "professional": "specialist",
    "senior": "senior",
    "senior specialist": "senior",
    "sr specialist": "senior",
    "manager": "manager",
    "teamlead": "manager",
    "team lead": "manager",
    "leidinggevende": "manager",
    "director": "director",
    "directeur": "director",
    "head": "director",
    "c_level": "c_level",
    "c-level": "c_level",
    "c level": "c_level",
    "directie": "c_level",
    "mt": "c_level",
}

BOUNDED_COMMERCE_CORE_ROUTE_INTERESTS = {"exitscan", "retentiescan"}
QUALIFICATION_CONFIRMABLE_ROUTE_INTERESTS = {
    "exitscan",
    "retentiescan",
    "combinatie",
    "teamscan",
    "onboarding",
    "leadership",
}


def _get_client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    real_ip = request.headers.get("x-real-ip")
    if real_ip:
        return real_ip.strip()
    return request.client.host if request.client else "unknown"


def _is_contact_rate_limited(ip: str) -> bool:
    now = time()
    bucket = _contact_request_buckets.get(ip)

    if not bucket or now > float(bucket["reset_at"]):
        _contact_request_buckets[ip] = {"count": 1, "reset_at": now + CONTACT_RATE_LIMIT_WINDOW_SECONDS}
        return False

    bucket["count"] = int(bucket["count"]) + 1
    return int(bucket["count"]) > CONTACT_RATE_LIMIT_MAX_REQUESTS


def _cleanup_contact_rate_limits() -> None:
    now = time()
    expired = [ip for ip, bucket in _contact_request_buckets.items() if now > float(bucket["reset_at"])]
    for ip in expired:
        _contact_request_buckets.pop(ip, None)


def _supports_bounded_commerce_route(route_interest: str | None) -> bool:
    return route_interest in BOUNDED_COMMERCE_CORE_ROUTE_INTERESTS


def _supports_confirmable_qualified_route(route_interest: str | None) -> bool:
    return route_interest in QUALIFICATION_CONFIRMABLE_ROUTE_INTERESTS


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _normalize_import_header(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip().lower()
    if not raw:
        return None
    raw = raw.replace("-", "_")
    return RESPONDENT_IMPORT_COLUMN_ALIASES.get(raw, RESPONDENT_IMPORT_COLUMN_ALIASES.get(raw.replace("_", " ")))


def _normalize_role_level(value: Any) -> str | None:
    if value is None:
        return None
    raw = str(value).strip().lower()
    if not raw:
        return None
    return ROLE_LEVEL_ALIASES.get(raw)


def _normalize_exit_month(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("/", "-").replace(".", "-")
    if len(raw) == 7 and raw[4] == "-":
        return raw
    return None


def _read_uploaded_rows(upload: UploadFile, content: bytes) -> list[tuple[int, dict[str, Any]]]:
    filename = (upload.filename or "").lower()
    if filename.endswith(".csv"):
        decoded: str
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(decoded))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV mist een headerregel met kolomnamen.")

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in enumerate(reader, start=2):
            rows.append((row_number, dict(row)))
        return rows

    if filename.endswith(".xlsx"):
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Excelbestand kon niet worden gelezen.") from exc

        worksheet = workbook[workbook.sheetnames[0]]
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            raise HTTPException(status_code=400, detail="Excelbestand mist een headerregel met kolomnamen.")

        headers = list(header_cells)
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            row = {str(headers[idx]): values[idx] for idx in range(len(headers)) if headers[idx] is not None}
            rows.append((row_number, row))
        return rows

    raise HTTPException(status_code=400, detail="Upload een .csv of .xlsx bestand.")


def _build_import_preview(
    *,
    raw_rows: list[tuple[int, dict[str, Any]]],
    existing_emails: set[str],
) -> tuple[list[RespondentCreate], list[RespondentImportIssue], list[RespondentImportPreviewRow], int]:
    issues: list[RespondentImportIssue] = []
    preview_rows: list[RespondentImportPreviewRow] = []
    normalized_rows: list[RespondentCreate] = []
    seen_emails: set[str] = set()
    duplicate_existing = 0

    for row_number, row in raw_rows:
        normalized = {
            canonical: None
            for canonical in ("email", "department", "role_level", "exit_month", "annual_salary_eur")
        }
        for key, value in row.items():
            canonical = _normalize_import_header(key)
            if not canonical:
                continue
            normalized[canonical] = value

        email = str(normalized["email"] or "").strip().lower()
        if not email:
            issues.append(RespondentImportIssue(row_number=row_number, field="email", message="E-mailadres ontbreekt."))
            continue

        department_raw = normalized["department"]
        department = str(department_raw).strip() if department_raw not in (None, "") else None

        role_level = _normalize_role_level(normalized["role_level"])
        if normalized["role_level"] not in (None, "") and not role_level:
            issues.append(
                RespondentImportIssue(
                    row_number=row_number,
                    field="role_level",
                    message="Functieniveau niet herkend. Gebruik uitvoerend, specialist, senior, manager, director of c_level.",
                )
            )
            continue

        exit_month = _normalize_exit_month(normalized["exit_month"])
        if normalized["exit_month"] not in (None, "") and not exit_month:
            issues.append(
                RespondentImportIssue(
                    row_number=row_number,
                    field="exit_month",
                    message="Gebruik voor exitmaand het formaat JJJJ-MM, bijvoorbeeld 2026-03.",
                )
            )
            continue

        salary_value = normalized["annual_salary_eur"]
        annual_salary_eur: float | None = None
        if salary_value not in (None, ""):
            try:
                annual_salary_eur = float(str(salary_value).replace(",", "."))
            except ValueError:
                issues.append(
                    RespondentImportIssue(
                        row_number=row_number,
                        field="annual_salary_eur",
                        message="Jaarsalaris moet numeriek zijn.",
                    )
                )
                continue

        if email in seen_emails:
            issues.append(
                RespondentImportIssue(
                    row_number=row_number,
                    field="email",
                    message="Dit e-mailadres staat dubbel in het bestand.",
                )
            )
            continue

        if email in existing_emails:
            duplicate_existing += 1
            issues.append(
                RespondentImportIssue(
                    row_number=row_number,
                    field="email",
                    message="Deze respondent bestaat al in deze campagne.",
                )
            )
            continue

        try:
            respondent = RespondentCreate(
                email=email,
                department=department,
                role_level=role_level,
                exit_month=exit_month,
                annual_salary_eur=annual_salary_eur,
            )
        except Exception as exc:
            issues.append(
                RespondentImportIssue(
                    row_number=row_number,
                    field="row",
                    message=str(exc),
                )
            )
            continue

        seen_emails.add(email)
        normalized_rows.append(respondent)
        if len(preview_rows) < 10:
            preview_rows.append(
                RespondentImportPreviewRow(
                    row_number=row_number,
                    email=email,
                    department=department,
                    role_level=role_level,
                    exit_month=exit_month,
                    annual_salary_eur=annual_salary_eur,
                )
            )

    return normalized_rows, issues, preview_rows, duplicate_existing


def _render_survey_status(
    request: Request,
    *,
    status_code: int,
    title: str,
    heading: str,
    message: str,
    hint: str | None = None,
    tone: str = "info",
) -> HTMLResponse:
    return templates.TemplateResponse(
        request,
        "survey-status.html",
        context={
            "title": title,
            "heading": heading,
            "message": message,
            "hint": hint,
            "tone": tone,
        },
        status_code=status_code,
    )


# ---------------------------------------------------------------------------
# Global DB error handler — geeft 503 terug bij verbindingsfouten
# (Supabase circuit breaker, network timeout, etc.) in plaats van HTTP 500
# ---------------------------------------------------------------------------

@app.exception_handler(OperationalError)
async def db_operational_error_handler(request: Request, exc: OperationalError):
    return JSONResponse(
        status_code=503,
        content={"detail": "Database tijdelijk niet beschikbaar. Probeer het over enkele seconden opnieuw."},
    )


@app.exception_handler(SQLAlchemyError)
async def db_general_error_handler(request: Request, exc: SQLAlchemyError):
    return JSONResponse(
        status_code=503,
        content={"detail": "Database-fout opgetreden. Probeer het opnieuw."},
    )


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------

@app.get("/api/health")
async def health() -> JSONResponse:
    db_ok = check_db_connection()
    return JSONResponse(
        status_code=200,
        content={"status": "ok" if db_ok else "degraded", "checks": {"database": db_ok}},
    )


@app.get("/api/ready")
async def ready() -> JSONResponse:
    db_ok = check_db_connection()
    return JSONResponse(
        status_code=200 if db_ok else 503,
        content={"status": "ok" if db_ok else "degraded", "checks": {"database": db_ok}},
    )


@app.post("/api/contact-request", response_model=ContactRequestResponse)
async def create_contact_request(
    body: ContactRequestCreate,
    request: Request,
    db: Session = Depends(get_db),
) -> ContactRequestResponse:
    if body.website:
        return ContactRequestResponse(message="Verstuurd", notification_sent=True)

    _cleanup_contact_rate_limits()
    client_ip = _get_client_ip(request)
    if _is_contact_rate_limited(client_ip):
        return JSONResponse(
            status_code=429,
            content={"detail": "Te veel aanvragen in korte tijd. Probeer het over 15 minuten opnieuw."},
        )

    lead = ContactRequest(
        name=body.name,
        work_email=body.work_email,
        organization=body.organization,
        employee_count=body.employee_count,
        route_interest=body.route_interest,
        cta_source=body.cta_source,
        desired_timing=body.desired_timing,
        current_question=body.current_question,
        website=body.website,
        notification_sent=False,
        notification_error=None,
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)

    send_result = send_contact_request_result(
        name=body.name,
        work_email=body.work_email,
        organization=body.organization,
        employee_count=body.employee_count,
        route_interest=body.route_interest,
        cta_source=body.cta_source,
        desired_timing=body.desired_timing,
        current_question=body.current_question,
    )

    if send_result.ok:
        lead.notification_sent = True
        lead.notification_error = None
        db.add(lead)
        db.commit()
        return ContactRequestResponse(
            message="Verstuurd",
            notification_sent=True,
            lead_id=lead.id,
        )
    else:
        lead.notification_error = send_result.reason
        db.add(lead)
        db.commit()
        logger.warning(
            "Contactaanvraag opgeslagen zonder e-mailnotificatie voor %s (%s): %s",
            body.organization,
            body.work_email,
            send_result.reason or "onbekende fout",
        )
        return JSONResponse(
            status_code=202,
            content=ContactRequestResponse(
                message="Aanvraag opgeslagen",
                notification_sent=False,
                warning=(
                    "Je aanvraag is opgeslagen, maar de e-mailnotificatie liep vast. "
                    "Controleer de leadlijst of de backendlogs voor verdere opvolging."
                ),
                lead_id=lead.id,
            ).model_dump(),
        )


@app.get("/api/contact-requests", response_model=list[ContactRequestRead])
async def list_contact_requests(
    limit: int = Query(default=25, ge=1, le=100),
    x_admin_token: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> list[ContactRequest]:
    require_backend_admin_token(x_admin_token, is_production=_IS_PRODUCTION)
    return (
        db.query(ContactRequest)
        .order_by(ContactRequest.created_at.desc())
        .limit(limit)
        .all()
    )


@app.patch("/api/contact-requests/{lead_id}", response_model=ContactRequestRead)
async def update_contact_request(
    lead_id: str,
    body: ContactRequestUpdate,
    x_admin_token: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> ContactRequest:
    require_backend_admin_token(x_admin_token, is_production=_IS_PRODUCTION)

    lead = db.query(ContactRequest).filter(ContactRequest.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Contactaanvraag niet gevonden.")

    if body.ops_stage is not None:
        lead.ops_stage = body.ops_stage
    if body.ops_exception_status is not None:
        lead.ops_exception_status = body.ops_exception_status
    if body.ops_owner is not None:
        lead.ops_owner = body.ops_owner.strip() or None
    if body.ops_next_step is not None:
        lead.ops_next_step = body.ops_next_step.strip() or None
    if body.ops_handoff_note is not None:
        lead.ops_handoff_note = body.ops_handoff_note.strip() or None

    qualification_update_requested = any(
        value is not None
        for value in (
            body.qualification_status,
            body.qualified_route,
            body.qualification_note,
            body.qualification_reviewed_by,
        )
    )

    qualification_status = body.qualification_status or lead.qualification_status
    qualified_route = (
        _normalize_optional_text(body.qualified_route)
        if body.qualified_route is not None
        else lead.qualified_route
    )
    qualification_note = (
        _normalize_optional_text(body.qualification_note)
        if body.qualification_note is not None
        else lead.qualification_note
    )
    qualification_reviewed_by = (
        _normalize_optional_text(body.qualification_reviewed_by)
        if body.qualification_reviewed_by is not None
        else lead.qualification_reviewed_by
    )
    qualification_reviewed_at = lead.qualification_reviewed_at
    now = datetime.now(timezone.utc)

    if qualification_update_requested:
        if qualification_status == "not_reviewed":
            qualified_route = None
            qualification_note = None
            qualification_reviewed_by = None
            qualification_reviewed_at = None
        elif qualification_status == "needs_route_review":
            qualified_route = None
            if body.qualification_reviewed_by is not None:
                qualification_reviewed_at = now if qualification_reviewed_by else None
            elif qualification_reviewed_by and qualification_reviewed_at is None:
                qualification_reviewed_at = now
        else:
            if not qualified_route or not _supports_confirmable_qualified_route(qualified_route):
                raise HTTPException(
                    status_code=422,
                    detail="Een bevestigde qualification vereist een expliciete gekwalificeerde route.",
                )
            if not qualification_reviewed_by:
                raise HTTPException(
                    status_code=422,
                    detail="Een bevestigde qualification vereist een expliciete reviewer.",
                )
            if body.qualification_reviewed_by is not None:
                qualification_reviewed_at = now if qualification_reviewed_by else None
            elif qualification_reviewed_by and qualification_reviewed_at is None:
                qualification_reviewed_at = now

        lead.qualification_status = qualification_status
        lead.qualified_route = qualified_route
        lead.qualification_note = qualification_note
        lead.qualification_reviewed_by = qualification_reviewed_by
        lead.qualification_reviewed_at = qualification_reviewed_at

    commerce_update_requested = any(
        value is not None
        for value in (
            body.commercial_agreement_status,
            body.commercial_pricing_mode,
            body.commercial_start_readiness_status,
            body.commercial_start_blocker,
            body.commercial_agreement_confirmed_by,
            body.commercial_readiness_reviewed_by,
        )
    )

    if commerce_update_requested and not _supports_bounded_commerce_route(lead.route_interest):
        raise HTTPException(
            status_code=422,
            detail="Bounded billing/check-out foundation ondersteunt in deze wave alleen ExitScan en RetentieScan.",
        )

    if commerce_update_requested:
        agreement_status = body.commercial_agreement_status or lead.commercial_agreement_status
        pricing_mode = (
            _normalize_optional_text(body.commercial_pricing_mode)
            if body.commercial_pricing_mode is not None
            else lead.commercial_pricing_mode
        )
        start_readiness_status = (
            body.commercial_start_readiness_status or lead.commercial_start_readiness_status
        )
        start_blocker = (
            _normalize_optional_text(body.commercial_start_blocker)
            if body.commercial_start_blocker is not None
            else lead.commercial_start_blocker
        )
        agreement_confirmed_by = (
            _normalize_optional_text(body.commercial_agreement_confirmed_by)
            if body.commercial_agreement_confirmed_by is not None
            else lead.commercial_agreement_confirmed_by
        )
        readiness_reviewed_by = (
            _normalize_optional_text(body.commercial_readiness_reviewed_by)
            if body.commercial_readiness_reviewed_by is not None
            else lead.commercial_readiness_reviewed_by
        )
        agreement_confirmed_at = lead.commercial_agreement_confirmed_at
        readiness_reviewed_at = lead.commercial_readiness_reviewed_at
        if agreement_status == "not_started":
            pricing_mode = None
            start_readiness_status = "not_ready"
            start_blocker = None
            agreement_confirmed_by = None
            agreement_confirmed_at = None
            readiness_reviewed_by = None
            readiness_reviewed_at = None
        elif agreement_status == "confirmed":
            if pricing_mode is None:
                raise HTTPException(
                    status_code=422,
                    detail="Bevestigd commercieel akkoord vereist een prijsmodus: public_anchor of custom_quote.",
                )
            if start_readiness_status == "blocked" and not start_blocker:
                raise HTTPException(
                    status_code=422,
                    detail="Een geblokkeerde start readiness vereist een expliciete blocker.",
                )
            if start_readiness_status == "ready" and not agreement_confirmed_by:
                raise HTTPException(
                    status_code=422,
                    detail="Start readiness kan pas op ready staan nadat intern akkoord expliciet is bevestigd.",
                )
            if start_readiness_status in {"ready", "blocked"} and not readiness_reviewed_by:
                raise HTTPException(
                    status_code=422,
                    detail="Delivery-start governance vereist een expliciete readiness review voor ready of blocked.",
                )
        else:
            if start_readiness_status == "ready":
                raise HTTPException(
                    status_code=422,
                    detail="Start readiness kan alleen op ready staan nadat commercieel akkoord bevestigd is.",
                )
            if start_readiness_status == "blocked" and not start_blocker:
                raise HTTPException(
                    status_code=422,
                    detail="Een geblokkeerde start readiness vereist een expliciete blocker.",
                )
            if start_readiness_status == "blocked" and not readiness_reviewed_by:
                raise HTTPException(
                    status_code=422,
                    detail="Delivery-start governance vereist een expliciete readiness review voor ready of blocked.",
                )
            agreement_confirmed_by = None
            agreement_confirmed_at = None

        if start_readiness_status != "blocked":
            start_blocker = None
        if start_readiness_status == "not_ready":
            readiness_reviewed_by = None
            readiness_reviewed_at = None

        if agreement_status == "confirmed":
            if body.commercial_agreement_confirmed_by is not None:
                agreement_confirmed_at = now if agreement_confirmed_by else None
            elif agreement_confirmed_by and agreement_confirmed_at is None:
                agreement_confirmed_at = now

        if start_readiness_status in {"ready", "blocked"}:
            if body.commercial_readiness_reviewed_by is not None:
                readiness_reviewed_at = now if readiness_reviewed_by else None
            elif readiness_reviewed_by and readiness_reviewed_at is None:
                readiness_reviewed_at = now

        lead.commercial_agreement_status = agreement_status
        lead.commercial_pricing_mode = pricing_mode
        lead.commercial_start_readiness_status = start_readiness_status
        lead.commercial_start_blocker = start_blocker
        lead.commercial_agreement_confirmed_by = agreement_confirmed_by
        lead.commercial_agreement_confirmed_at = agreement_confirmed_at
        lead.commercial_readiness_reviewed_by = readiness_reviewed_by
        lead.commercial_readiness_reviewed_at = readiness_reviewed_at

    if lead.ops_stage == "implementation_intake_ready":
        if lead.qualification_status != "route_confirmed" or not lead.qualified_route or not lead.qualification_reviewed_by:
            raise HTTPException(
                status_code=422,
                detail="Implementation intake ready vereist eerst een bevestigde qualificationroute en expliciete qualification review.",
            )

    if body.last_contacted_at is not None:
        lead.last_contacted_at = body.last_contacted_at

    db.add(lead)
    db.commit()
    db.refresh(lead)
    return lead


@app.get("/survey/complete", response_class=HTMLResponse)
async def survey_complete(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "complete.html", context={"already_done": False})


# ---------------------------------------------------------------------------
# Survey — serve HTML form
# ---------------------------------------------------------------------------

@app.get("/survey/{token}", response_class=HTMLResponse)
async def serve_survey(
    request: Request,
    token: str,
    db: Session = Depends(get_db),
) -> HTMLResponse:
    respondent = _get_respondent_by_token(token, db)

    if not respondent:
        return _render_survey_status(
            request,
            status_code=404,
            title="Survey-link ongeldig",
            heading="Deze survey-link klopt niet",
            message="De link is ongeldig of niet meer actief. Controleer of je de volledige link hebt geopend.",
            hint="Neem contact op met de HR-afdeling van je organisatie als je een nieuwe uitnodiging nodig hebt.",
            tone="warning",
        )

    # Controleer of token verlopen is (90 dagen geldig)
    if hasattr(respondent, "token_expires_at") and respondent.token_expires_at:
        if respondent.token_expires_at < datetime.now(timezone.utc):
            return _render_survey_status(
                request,
                status_code=410,
                title="Survey-link verlopen",
                heading="Deze survey-link is verlopen",
                message="Om privacy- en beveiligingsredenen is deze persoonlijke link niet meer geldig.",
                hint="Neem contact op met de HR-afdeling van je organisatie als je de survey alsnog wilt invullen.",
                tone="warning",
            )

    if respondent.completed:
        return templates.TemplateResponse(
            request,
            "complete.html",
            context={"already_done": True},
        )

    # Mark opened
    if not respondent.opened_at:
        respondent.opened_at = datetime.now(timezone.utc)
        db.commit()

    campaign = respondent.campaign
    if not campaign.is_active:
        return _render_survey_status(
            request,
            status_code=410,
            title="Survey gesloten",
            heading="Deze survey is gesloten",
            message="Deze campagne accepteert geen nieuwe inzendingen meer. Eerder ingevulde antwoorden blijven wel meegenomen in de rapportage.",
            hint="Heb je vragen over de uitkomsten of verwerking van je gegevens, neem dan contact op met de HR-afdeling van je organisatie.",
            tone="info",
        )

    enabled_modules = _resolve_survey_modules(campaign.scan_type, campaign.enabled_modules)
    scan_definition = get_scan_definition(campaign.scan_type)

    return templates.TemplateResponse(
        request,
        "survey.html",
        context={
            "token":           token,
            "scan":            scan_definition,
            "scan_type":       campaign.scan_type,
            "campaign_name":   campaign.name,
            "enabled_modules": enabled_modules,
        },
    )


# ---------------------------------------------------------------------------
# Survey — process submission
# ---------------------------------------------------------------------------

@app.post("/survey/submit", response_class=JSONResponse)
async def submit_survey(
    payload: SurveySubmit,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    # Validate token
    respondent = _get_respondent_by_token(payload.token, db)
    if not respondent:
        raise HTTPException(status_code=404, detail="Ongeldige token.")
    if respondent.completed:
        raise HTTPException(status_code=409, detail="Survey al ingevuld.")
    if not respondent.campaign.is_active:
        raise HTTPException(status_code=410, detail="Deze survey is gesloten en accepteert geen nieuwe inzendingen meer.")
    product_module = get_product_module(respondent.campaign.scan_type)
    product_module.validate_submission(payload)

    exit_reason_code = payload.exit_reason_code
    if not exit_reason_code and payload.exit_reason_category:
        exit_reason_code = EXIT_REASON_CODE_MAP.get(payload.exit_reason_category)
    contributing_reason_codes = [
        mapped_code
        for mapped_code in (
            EXIT_REASON_CODE_MAP.get(reason_key, reason_key)
            for reason_key in payload.pull_factors_raw.keys()
        )
        if mapped_code and mapped_code != exit_reason_code
    ]
    normalized_contributing_reasons = {code: 1 for code in contributing_reason_codes}

    # --- Scoring ---
    try:
        score_payload = product_module.score_submission(
            payload=payload,
            campaign=respondent.campaign,
            respondent=respondent,
            exit_reason_code=exit_reason_code,
            contributing_reason_codes=contributing_reason_codes,
        )
        sdt_scores = score_payload["sdt_scores"]
        org_scores = score_payload["org_scores"]
        risk_result = score_payload["risk_result"]
        preventability_result = score_payload["preventability_result"]
        replacement_cost_eur = score_payload["replacement_cost_eur"]
        recommendations = score_payload["recommendations"]
        uwes_score = score_payload["uwes_score"]
        ti_score = score_payload["turnover_intention_score"]
        retention_summary = score_payload["retention_summary"]
        full_result = score_payload["full_result"]
        scoring_version = score_payload["scoring_version"]
    except Exception as exc:
        sentry_sdk.capture_exception(exc, extras={
            "campaign_id": str(respondent.campaign_id),
            "flow": "survey_scoring",
        })
        raise HTTPException(status_code=500, detail="Scoringsberekening mislukt. De survey is niet opgeslagen.")

    # Open text — anonymise then optionally enrich
    open_text_clean: str | None = None
    open_text_analysis: str | None = None
    if payload.open_text:
        open_text_clean = anonymize_text(payload.open_text)
        if _openai_available and open_text_clean:
            open_text_analysis = _enrich_open_text(open_text_clean)


    # --- Persist ---
    response_row = SurveyResponse(
        respondent_id         = respondent.id,
        tenure_years          = payload.tenure_years,
        exit_reason_category  = payload.exit_reason_category,
        exit_reason_code      = exit_reason_code,
        stay_intent_score     = payload.stay_intent_score,
        sdt_raw               = payload.sdt_raw,
        sdt_scores            = sdt_scores,
        org_raw               = payload.org_raw,
        org_scores            = org_scores,
        pull_factors_raw      = normalized_contributing_reasons,
        open_text_raw         = open_text_clean,
        open_text_analysis    = open_text_analysis,
        uwes_raw              = payload.uwes_raw,
        uwes_score            = uwes_score,
        turnover_intention_raw   = payload.turnover_intention_raw,
        turnover_intention_score = ti_score,
        risk_score            = risk_result["risk_score"],
        risk_band             = risk_result["risk_band"],
        preventability        = preventability_result.get("preventability"),
        replacement_cost_eur  = replacement_cost_eur,
        full_result           = full_result,
        scoring_version       = scoring_version,
    )

    # AVG data minimalisatie: jaarlijks salaris is niet langer nodig na scoring.
    # De replacement_cost_eur is berekend en opgeslagen; het bronsalaris wordt verwijderd.
    if respondent.annual_salary_eur is not None:
        respondent.annual_salary_eur = None

    db.add(response_row)

    respondent.completed    = True
    respondent.completed_at = datetime.now(timezone.utc)
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        sentry_sdk.capture_exception(exc, extras={
            "campaign_id": str(respondent.campaign_id),
            "flow": "survey_persist",
        })
        raise HTTPException(status_code=500, detail="Opslaan van survey-antwoorden mislukt. Probeer opnieuw.")

    campaign_obj = respondent.campaign
    org = campaign_obj.organization
    total_completed = sum(1 for r in campaign_obj.respondents if r.completed)
    total_invited = len(campaign_obj.respondents)
    background_tasks.add_task(
        _send_hr_notification_safe,
        to_email=org.contact_email,
        campaign_name=campaign_obj.name,
        campaign_id=campaign_obj.id,
        total_completed=total_completed,
        total_invited=total_invited,
    )

    return {"status": "ok", "message": "Bedankt voor het invullen van de survey."}


# ---------------------------------------------------------------------------
# Organisation management
# ---------------------------------------------------------------------------

@app.post("/api/organizations", response_model=OrganizationRead, status_code=201)
async def create_organization(
    body: OrganizationCreate,
    x_admin_token: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> Organization:
    require_backend_admin_token(x_admin_token, is_production=_IS_PRODUCTION)

    existing = db.query(Organization).filter(Organization.slug == body.slug).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Slug '{body.slug}' al in gebruik.")

    bind = db.get_bind()
    is_sqlite_session = bind is not None and bind.dialect.name == "sqlite"

    if not is_sqlite_session:
        admin_profile_id = db.execute(
            text(
                """
                select id
                from public.profiles
                where is_verisight_admin = true
                order by created_at asc
                limit 1
                """
            )
        ).scalar()
        if not admin_profile_id:
            raise HTTPException(
                status_code=503,
                detail="Geen Verisight admin-profiel beschikbaar om een organisatie aan te maken.",
            )

        # Supabase-triggers gebruiken auth.uid() om org ownership te registreren.
        db.execute(
            text("select set_config('request.jwt.claim.sub', :user_id, true)"),
            {"user_id": str(admin_profile_id)},
        )

    org = Organization(
        name=body.name,
        slug=body.slug,
        contact_email=body.contact_email,
    )
    db.add(org)
    db.flush()

    # In Postgres/Supabase kan een DB-trigger het organization secret al aanmaken.
    # Alleen lokaal aanvullen als er na de org-insert nog geen secret bestaat.
    existing_secret = (
        db.query(OrganizationSecret)
        .filter(OrganizationSecret.org_id == org.id)
        .first()
    )
    if not existing_secret:
        db.add(OrganizationSecret(org_id=org.id))

    db.commit()
    db.refresh(org)
    return org


@app.get("/api/organizations/{slug}", response_model=OrganizationRead)
async def get_organization(slug: str, db: Session = Depends(get_db)) -> Organization:
    org = db.query(Organization).filter(Organization.slug == slug).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organisatie niet gevonden.")
    return org


# ---------------------------------------------------------------------------
# Campaign management
# ---------------------------------------------------------------------------

def _get_org_by_api_key(api_key: str, db: Session) -> Organization:
    secret = (
        db.query(OrganizationSecret)
        .options(selectinload(OrganizationSecret.organization))
        .filter(OrganizationSecret.api_key == api_key)
        .first()
    )
    org = secret.organization if secret else None

    # Backwards-compatible fallback while older databases still store api_key on organizations.
    if not org:
        try:
            row = db.execute(
                text(
                    """
                    select id
                    from organizations
                    where api_key = :api_key
                    limit 1
                    """
                ),
                {"api_key": api_key},
            ).first()
            if row:
                org = db.query(Organization).filter(Organization.id == str(row[0])).first()
        except Exception:
            org = None

    if not org:
        raise HTTPException(status_code=401, detail="Ongeldige API-sleutel.")
    if not org.is_active:
        raise HTTPException(status_code=403, detail="Organisatie is inactief.")
    return org


def _get_respondent_by_token(token: str, db: Session) -> Respondent | None:
    try:
        normalized = str(UUID(token))
    except (TypeError, ValueError):
        return None

    return db.query(Respondent).filter(Respondent.token == normalized).first()


def _create_campaign_respondents(
    *,
    campaign: Campaign,
    respondent_inputs: list[RespondentCreate],
    db: Session,
    send_invites: bool,
) -> tuple[list[Respondent], int]:
    respondents = [
        Respondent(
            campaign_id=campaign.id,
            department=r.department,
            role_level=r.role_level,
            exit_month=r.exit_month,
            annual_salary_eur=r.annual_salary_eur,
            email=r.email,
        )
        for r in respondent_inputs
    ]
    db.add_all(respondents)
    db.flush()
    db.commit()

    emails_sent = 0
    if send_invites:
        for respondent, req in zip(respondents, respondent_inputs):
            if req.email:
                sent = send_survey_invite(
                    to_email=req.email,
                    campaign_name=campaign.name,
                    token=respondent.token,
                    scan_type=campaign.scan_type,
                )
                if sent:
                    respondent.sent_at = datetime.now(timezone.utc)
                    emails_sent += 1
        db.commit()

    return respondents, emails_sent


@app.post("/api/campaigns", response_model=CampaignRead, status_code=201)
async def create_campaign(
    body: CampaignCreate,
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> Campaign:
    org = _get_org_by_api_key(x_api_key, db)
    campaign = Campaign(
        organization_id=org.id,
        name=body.name,
        scan_type=body.scan_type,
        delivery_mode=body.delivery_mode or "baseline",
        enabled_modules=body.enabled_modules,
    )
    db.add(campaign)
    db.commit()
    db.refresh(campaign)
    return campaign


@app.get("/api/campaigns", response_model=list[CampaignRead])
async def list_campaigns(
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> list[Campaign]:
    org = _get_org_by_api_key(x_api_key, db)
    return db.query(Campaign).filter(Campaign.organization_id == org.id).all()


# ---------------------------------------------------------------------------
# Respondent management
# ---------------------------------------------------------------------------

@app.post("/api/campaigns/{campaign_id}/respondents", response_model=list[RespondentRead], status_code=201)
async def add_respondents(
    campaign_id: str,
    body: RespondentBatchCreate,
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> list[Respondent]:
    org = _get_org_by_api_key(x_api_key, db)
    campaign = (
        db.query(Campaign)
        .options(selectinload(Campaign.respondents).selectinload(Respondent.response))
        .filter(
            Campaign.id == campaign_id,
            Campaign.organization_id == org.id,
        )
        .first()
    )
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")
    respondents, _emails_sent = _create_campaign_respondents(
        campaign=campaign,
        respondent_inputs=body.respondents,
        db=db,
        send_invites=body.send_invites,
    )
    return respondents


@app.post("/api/campaigns/{campaign_id}/respondents/import", response_model=RespondentImportResponse)
async def import_respondents(
    campaign_id: str,
    x_api_key: Annotated[str, Header()],
    upload: UploadFile = File(...),
    dry_run: bool = Form(True),
    send_invites: bool = Form(True),
    db: Session = Depends(get_db),
) -> RespondentImportResponse:
    org = _get_org_by_api_key(x_api_key, db)
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.organization_id == org.id,
    ).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")

    content = await upload.read()
    if not content:
        raise HTTPException(status_code=400, detail="Het uploadbestand is leeg.")

    raw_rows = _read_uploaded_rows(upload, content)
    if len(raw_rows) == 0:
        raise HTTPException(status_code=400, detail="Geen respondenten gevonden in het bestand.")
    if len(raw_rows) > 500:
        raise HTTPException(status_code=400, detail="Upload maximaal 500 respondenten per bestand.")

    existing_emails = {
        (email or "").strip().lower()
        for (email,) in db.query(Respondent.email)
        .filter(Respondent.campaign_id == campaign_id, Respondent.email.isnot(None))
        .all()
        if email
    }
    normalized_rows, issues, preview_rows, duplicate_existing = _build_import_preview(
        raw_rows=raw_rows,
        existing_emails=existing_emails,
    )

    response = RespondentImportResponse(
        dry_run=dry_run,
        total_rows=len(raw_rows),
        valid_rows=len(normalized_rows),
        invalid_rows=len(issues),
        duplicate_existing=duplicate_existing,
        preview_rows=preview_rows,
        errors=issues,
        imported=0,
        emails_sent=0,
    )

    if dry_run or issues:
        return response

    respondents, emails_sent = _create_campaign_respondents(
        campaign=campaign,
        respondent_inputs=normalized_rows,
        db=db,
        send_invites=send_invites,
    )
    response.imported = len(respondents)
    response.emails_sent = emails_sent
    return response


@app.get("/api/campaigns/{campaign_id}/respondents", response_model=list[RespondentRead])
async def list_respondents(
    campaign_id: str,
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> list[Respondent]:
    org = _get_org_by_api_key(x_api_key, db)
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.organization_id == org.id,
    ).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")
    return campaign.respondents


@app.post("/api/campaigns/{campaign_id}/send-invites", response_model=InviteSendResult)
async def send_invites(
    campaign_id: str,
    body: list[SendInviteItem],
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> InviteSendResult:
    """
    Stuur uitnodigingsmails naar een lijst van {token, email} paren.
    Wordt aangeroepen vanuit het dashboard direct na het aanmaken van respondenten.
    Vereist een geldige organisatie-API-key; opgeslagen respondentdata blijft leidend.
    """
    org = _get_org_by_api_key(x_api_key, db)
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.organization_id == org.id,
    ).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")

    sent = 0
    failed = 0
    skipped = 0

    requested_tokens = [item.token for item in body]
    respondents = (
        db.query(Respondent)
        .filter(
            Respondent.campaign_id == campaign_id,
            Respondent.token.in_(requested_tokens),
        )
        .all()
    )
    respondents_by_token = {respondent.token: respondent for respondent in respondents}

    for item in body:
        respondent = respondents_by_token.get(item.token)

        if not respondent or not respondent.email:
            failed += 1
            continue

        if item.email and item.email != respondent.email:
            skipped += 1
            continue

        ok = send_survey_invite(
            to_email=respondent.email,
            campaign_name=campaign.name,
            token=item.token,
            scan_type=campaign.scan_type,
        )

        if ok:
            respondent.sent_at = datetime.now(timezone.utc)
            sent += 1
        else:
            failed += 1

    if sent > 0:
        db.commit()

    if failed > 0 and sent == 0:
        import logging
        logging.getLogger(__name__).warning(
            "Alle %s uitnodigingen mislukt voor campaign %s",
            failed,
            campaign_id,
        )

    return InviteSendResult(sent=sent, failed=failed, skipped=skipped)


# ---------------------------------------------------------------------------
# Campaign lifecycle management
# ---------------------------------------------------------------------------
# Sluit-actie en herinnerings-actie zijn verplaatst naar Next.js server actions.
# Sluiten gaat direct via Supabase (RLS bewaakt schrijftoegang).
# Heruitzenden gaat via /api/campaigns/{id}/send-invites na auth-check in server action.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Campaign stats / dashboard data
# ---------------------------------------------------------------------------

@app.get("/api/campaigns/{campaign_id}/stats", response_model=CampaignStats)
async def campaign_stats(
    campaign_id: str,
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    org = _get_org_by_api_key(x_api_key, db)
    campaign = (
        db.query(Campaign)
        .options(selectinload(Campaign.respondents).selectinload(Respondent.response))
        .filter(
            Campaign.id == campaign_id,
            Campaign.organization_id == org.id,
        )
        .first()
    )
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")

    respondents = campaign.respondents
    completed = [respondent for respondent in respondents if respondent.completed and respondent.response]
    total       = len(respondents)
    n_completed = len(completed)

    responses: list[SurveyResponse] = [respondent.response for respondent in completed if respondent.response]

    # Risk distribution
    band_dist: dict[str, int] = {"HOOG": 0, "MIDDEN": 0, "LAAG": 0}
    risk_scores: list[float]  = []
    for resp in responses:
        if resp.risk_band in band_dist:
            band_dist[resp.risk_band] += 1
        if resp.risk_score is not None:
            risk_scores.append(resp.risk_score)

    avg_risk = round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else None

    # Pattern detection input
    pattern_input = [
        {
            "org_scores": respondent.response.org_scores,
            "sdt_scores": respondent.response.sdt_scores,
            "risk_score": respondent.response.risk_score,
            "preventability": respondent.response.preventability,
            "exit_reason_code": respondent.response.exit_reason_code,
            "contributing_reason_codes": list((respondent.response.pull_factors_raw or {}).keys()),
            "department": respondent.department,
            "role_level": respondent.role_level,
        }
        for respondent in completed
    ]
    pattern_report = detect_patterns(pattern_input) if pattern_input else None

    return {
        "campaign_id":              str(campaign.id),
        "campaign_name":            campaign.name,
        "organization_id":          str(campaign.organization_id),
        "scan_type":                campaign.scan_type,
        "is_active":                campaign.is_active,
        "created_at":               campaign.created_at,
        "total_invited":            total,
        "total_completed":          n_completed,
        "completion_rate_pct":      round(n_completed / total * 100, 1) if total > 0 else 0.0,
        "avg_risk_score":           avg_risk,
        "avg_signal_score":         avg_risk,
        "band_high":                band_dist["HOOG"],
        "band_medium":              band_dist["MIDDEN"],
        "band_low":                 band_dist["LAAG"],
        "risk_band_distribution":   band_dist,
        "pattern_report":           pattern_report,
    }


# ---------------------------------------------------------------------------
# PDF-rapport
# ---------------------------------------------------------------------------

@app.get("/api/campaigns/{campaign_id}/report")
async def download_report(
    campaign_id: str,
    x_api_key: Annotated[str, Header()],
    db: Session = Depends(get_db),
) -> Response:
    """Genereer en download een PDF-rapport voor de campaign."""
    org = _get_org_by_api_key(x_api_key, db)
    campaign = db.query(Campaign).filter(
        Campaign.id == campaign_id,
        Campaign.organization_id == org.id,
    ).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")
    if campaign.scan_type in {"pulse"}:
        product_name = (
            "Pulse"
            if campaign.scan_type == "pulse"
            else "TeamScan"
            if campaign.scan_type == "team"
            else "Leadership Scan"
        )
        raise HTTPException(status_code=422, detail=f"{product_name} ondersteunt in deze wave nog geen PDF-rapport.")

    try:
        from backend.report import generate_campaign_report
        pdf_bytes = generate_campaign_report(campaign_id, db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF-generatie mislukt: {e}")

    # HTTP Content-Disposition headers zijn latin-1; strip alle niet-ASCII tekens
    import re as _re
    safe_name = _re.sub(r"[^\w\s-]", "", campaign.name)   # verwijder em-dash etc.
    safe_name = _re.sub(r"[\s-]+", "_", safe_name).strip("_")  # spaties/streepjes → _
    filename = f"Verisight_{safe_name}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/internal/campaigns/{campaign_id}/report")
async def download_report_internal(
    campaign_id: str,
    x_admin_token: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> Response:
    """Interne rapportdownload voor vertrouwde server-side proxy's."""
    require_backend_admin_token(x_admin_token, is_production=_IS_PRODUCTION)

    campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign niet gevonden.")
    if campaign.scan_type in {"pulse"}:
        product_name = (
            "Pulse"
            if campaign.scan_type == "pulse"
            else "TeamScan"
            if campaign.scan_type == "team"
            else "Leadership Scan"
        )
        raise HTTPException(status_code=422, detail=f"{product_name} ondersteunt in deze wave nog geen PDF-rapport.")

    try:
        from backend.report import generate_campaign_report
        pdf_bytes = generate_campaign_report(campaign_id, db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF-generatie mislukt: {e}")

    import re as _re
    safe_name = _re.sub(r"[^\w\s-]", "", campaign.name)
    safe_name = _re.sub(r"[\s-]+", "_", safe_name).strip("_")
    filename = f"Verisight_{safe_name}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Open text enrichment (internal helper)
# ---------------------------------------------------------------------------

def _enrich_open_text(text: str) -> str | None:
    """Call OpenAI GPT-4o to extract themes from anonymised open text."""
    if not _openai_available or not _openai_client:
        return None

    prompt = (
        "Je bent een HR-analist. Analyseer de volgende anonieme toelichting van een "
        "medewerker en identificeer maximaal 3 concrete thema's of verbeterpunten. "
        "Geef een beknopte, professionele samenvatting in het Nederlands (max 100 woorden). "
        "Vermeld GEEN namen, persoonlijke gegevens of herleidbare informatie.\n\n"
        f"Toelichting:\n{text}"
    )

    for attempt in range(3):
        try:
            resp = _openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=200,
                temperature=0.3,
            )
            return resp.choices[0].message.content.strip()
        except Exception:
            import time
            time.sleep(2 ** attempt)

    return None
