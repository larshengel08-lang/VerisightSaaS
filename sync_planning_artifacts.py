from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import yaml


ROOT = Path(__file__).resolve().parent
DATA_PATH = ROOT / "docs" / "strategy" / "ROADMAP_DATA.yaml"
CHECKLIST_PATH = ROOT / "docs" / "prompts" / "PROMPT_CHECKLIST.xlsx"
ROADMAP_PATH = ROOT / "docs" / "strategy" / "ROADMAP.md"
ROADMAP_WORKBOOK_PATH = ROOT / "docs" / "strategy" / "ROADMAP_WORKBOOK.xlsx"


@dataclass
class ChecklistItem:
    prompt_file: str
    deliverable: str
    title: str
    phase: str
    default_note: str


def load_data() -> dict[str, Any]:
    return yaml.safe_load(DATA_PATH.read_text(encoding="utf-8"))


def load_existing_checklist_state() -> dict[str, list[Any]]:
    if not CHECKLIST_PATH.exists():
        return {}

    wb = openpyxl.load_workbook(CHECKLIST_PATH)
    ws = wb.active
    existing: dict[str, list[Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            values = list(row)
            if len(values) < 6:
                values += [None] * (6 - len(values))
            existing[str(values[0])] = values
    return existing


def build_checklist_rows(data: dict[str, Any], existing: dict[str, list[Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for raw_item in data["checklist"]["items"]:
        item = ChecklistItem(**raw_item)
        old = existing.get(item.prompt_file)
        if old:
            status = old[2] or "Niet voldaan"
            last_updated = old[3]
            notes = old[4]
            live = old[5]
            if status != "Voldaan":
                notes = item.default_note
            rows.append([item.prompt_file, item.deliverable, status, last_updated, notes, live])
        else:
            rows.append([item.prompt_file, item.deliverable, "Niet voldaan", None, item.default_note, None])
    return rows


def save_checklist(data: dict[str, Any], rows: list[list[Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"
    ws.append(data["checklist"]["header"])
    for row in rows:
        ws.append(row)
    wb.save(CHECKLIST_PATH)


def save_roadmap_workbook(data: dict[str, Any], rows: list[list[Any]]) -> None:
    wb = openpyxl.Workbook()

    overview = wb.active
    overview.title = "Overview"
    overview.append(["Field", "Value"])
    overview.append(["Title", f"{data['metadata']['title']} - Roadmap"])
    overview.append(["Generated", datetime.now(timezone.utc).date().isoformat()])
    overview.append(["Checklist Path", data["metadata"]["tactical_queue_path"]])
    overview.append(["Roadmap Path", data["metadata"]["roadmap_path"]])
    overview.append(["Strategy Path", data["metadata"]["strategy_path"]])
    overview.append(["External Docs Path", data["metadata"]["external_docs_path"]])
    overview.append([])
    overview.append(["Current State", ""])
    for bullet in data["metadata"]["current_state_intro"]:
        overview.append(["", bullet])

    post_checklist_system = data["metadata"].get("post_checklist_system")
    if post_checklist_system:
        overview.append([])
        overview.append([post_checklist_system["title"], ""])
        for bullet in post_checklist_system.get("bullets", []):
            overview.append(["", bullet])

    execution_priorities = data["metadata"].get("execution_priorities")
    if execution_priorities:
        overview.append([])
        overview.append(["Current Focus", ""])
        for item in execution_priorities.get("active_initiatives", []):
            overview.append(["", item["priority"]])
        for item in execution_priorities.get("next_30_days", []):
            overview.append(["", f"Next 30d: {item['priority']}"])

    phases_sheet = wb.create_sheet("Phases")
    phases_sheet.append(
        [
            "Phase ID",
            "Title",
            "Objective",
            "Status",
            "Checklist Items",
            "Exit Gate",
            "Waits Until Later",
        ]
    )
    rows_by_prompt = {str(row[0]): row for row in rows}
    phases = data["phases"]
    current_phase_id = next(
        (
            phase["id"]
            for phase in phases
            if any(rows_by_prompt.get(key, [None, None, "Niet voldaan"])[2] != "Voldaan" for key in phase["items"])
        ),
        phases[-1]["id"],
    )
    title_by_prompt = {item["prompt_file"]: item["title"] for item in data["checklist"]["items"]}
    for phase in phases:
        status_lines = phase_status(rows_by_prompt, phase["items"], current_phase_id, phase["id"])
        checklist_items = "\n".join(
            [
                f"{title_by_prompt[key]} - {'afgerond' if rows_by_prompt.get(key, [None, None, 'Niet voldaan'])[2] == 'Voldaan' else 'open'}"
                for key in phase["items"]
            ]
        )
        phases_sheet.append(
            [
                phase["id"],
                phase["title"],
                phase["objective"],
                " | ".join(status_lines).replace("- ", ""),
                checklist_items,
                "\n".join(phase["exit_gate"]),
                "\n".join(phase["waits_until_later"]),
            ]
        )

    checklist_sheet = wb.create_sheet("Checklist Mirror")
    checklist_sheet.append(
        [
            "Prompt File",
            "Deliverable",
            "Title",
            "Phase",
            "Status",
            "Last Updated",
            "Notes",
            "Repo/Main/Deploy/Live",
        ]
    )
    item_lookup = {item["prompt_file"]: item for item in data["checklist"]["items"]}
    for row in rows:
        prompt_file = str(row[0])
        item = item_lookup[prompt_file]
        checklist_sheet.append(
            [
                prompt_file,
                row[1],
                item["title"],
                item["phase"],
                row[2],
                row[3],
                row[4],
                row[5],
            ]
        )

    if execution_priorities:
        priorities_sheet = wb.create_sheet("Current Priorities")
        priorities_sheet.append(["Horizon", "Priority", "Why now", "Owner", "Status / guardrail"])

        def append_priority_block(label: str, items: list[dict[str, str]]) -> None:
            for item in items:
                priorities_sheet.append(
                    [
                        label,
                        item["priority"],
                        item["why"],
                        item["owner"],
                        "",
                    ]
                )

        append_priority_block("Active now", execution_priorities.get("active_initiatives", []))
        append_priority_block("Next 30 days", execution_priorities.get("next_30_days", []))
        append_priority_block("30-60 days", execution_priorities.get("next_30_60_days", []))
        append_priority_block("60-90 days", execution_priorities.get("next_60_90_days", []))

        priorities_sheet.append([])
        priorities_sheet.append(["Not now", "Guardrail", "", "", ""])
        for item in execution_priorities.get("not_now", []):
            priorities_sheet.append(["Not now", item, "", "", ""])

        weekly_focus = wb.create_sheet("Weekly Focus")
        weekly_focus.append(["Field", "Value", "Prompt"])
        weekly_focus.append(["Week van", "", "Kies de week die nu loopt."])
        weekly_focus.append(["Deze week telt vooral", "", "Welke ene beweging maakt deze week succesvol?"])
        weekly_focus.append(["Belangrijkste commerciële push", "", "Welk gesprek, voorstel of batch moet deze week echt bewegen?"])
        weekly_focus.append(["Belangrijkste delivery-risk", "", "Welk klanttraject vraagt nu de meeste aandacht?"])
        weekly_focus.append(["Topprioriteit 1", execution_priorities.get("active_initiatives", [{}])[0].get("priority", ""), "Houd dit concreet en uitvoerbaar."])
        weekly_focus.append(["Topprioriteit 2", execution_priorities.get("next_30_days", [{}])[0].get("priority", ""), "Kies een echte fix, geen abstract thema."])
        weekly_focus.append(["Topprioriteit 3", execution_priorities.get("next_30_days", [{}, {}])[1].get("priority", ""), "Alleen invullen als je het deze week echt doet."])
        weekly_focus.append(["Niet doen 1", "", "Wat laat je deze week bewust liggen?"])
        weekly_focus.append(["Niet doen 2", "", "Wat laat je deze week bewust liggen?"])
        weekly_focus.append(["Niet doen 3", "", "Wat laat je deze week bewust liggen?"])
        weekly_focus.append(["Besluit deze week", "", "Welk besluit moet expliciet in de decision log landen?"])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
                cell.alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=True)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)
        sheet.freeze_panes = "A2"

    wb.save(ROADMAP_WORKBOOK_PATH)


def release_status_warnings(rows: list[list[Any]]) -> list[str]:
    warnings: list[str] = []
    legacy_markers = (
        "live via main",
        "live op main",
        "uitgevoerd in repo",
    )
    for row in rows:
        prompt_file = str(row[0])
        status = "" if row[2] is None else str(row[2]).strip().lower()
        release_state = "" if row[5] is None else str(row[5]).strip()
        release_state_lower = release_state.lower()
        if status != "voldaan":
            continue
        if not release_state:
            warnings.append(f"{prompt_file}: status is Voldaan maar repo/main/deploy/live-status ontbreekt.")
            continue
        if any(marker in release_state_lower for marker in legacy_markers):
            warnings.append(
                f"{prompt_file}: gebruikt legacy release-status '{release_state}'. "
                "Vervang door expliciete repo-only of live-bevestigde formulering."
            )
    return warnings


def phase_status(rows_by_prompt: dict[str, list[Any]], items: list[str], current_phase_id: str, phase_id: str) -> list[str]:
    completed = sum(1 for key in items if rows_by_prompt.get(key, [None, None, "Niet voldaan"])[2] == "Voldaan")
    total = len(items)
    lines: list[str] = []
    if total == 0:
        lines.append("- gepland")
        return lines
    if completed == total:
        lines.append("- afgerond")
    elif phase_id == current_phase_id:
        lines.append("- huidige actieve focus")
        if completed > 0:
            lines.append(f"- {completed} van {total} checklist-items afgerond")
    elif completed > 0:
        lines.append(f"- deels uitgevoerd ({completed}/{total})")
    else:
        lines.append("- gepland")
    return lines


def build_roadmap_markdown(data: dict[str, Any], rows: list[list[Any]]) -> str:
    rows_by_prompt = {str(row[0]): row for row in rows}
    phases = data["phases"]

    current_phase_id = next(
        (
            phase["id"]
            for phase in phases
            if any(rows_by_prompt.get(key, [None, None, "Niet voldaan"])[2] != "Voldaan" for key in phase["items"])
        ),
        phases[-1]["id"],
    )

    lines: list[str] = []
    lines.append(f"# {data['metadata']['title']} - Roadmap")
    lines.append("")
    lines.append("Levend document. Dit bestand wordt gegenereerd uit `docs/strategy/ROADMAP_DATA.yaml` en de actuele status in `docs/prompts/PROMPT_CHECKLIST.xlsx`.")
    lines.append(f"Laatst gegenereerd: {datetime.now(timezone.utc).date().isoformat()}")
    lines.append("")
    lines.append("## Gebruik")
    lines.append("")
    lines.append(f"- [PROMPT_CHECKLIST.xlsx]({data['metadata']['tactical_queue_path']}) is de tactische uitvoerqueue.")
    lines.append("- Deze roadmap beschrijft de strategische fases, gates en afhankelijkheden.")
    lines.append(f"- Externe documenten in [{Path(data['metadata']['external_docs_path']).name}]({data['metadata']['external_docs_path']}) zijn waardevolle referentie, maar mogen geen concurrerende roadmap vormen.")
    lines.append("")
    lines.append("## Huidige stand")
    lines.append("")
    for bullet in data["metadata"]["current_state_intro"]:
        lines.append(f"- {bullet}")
    lines.append("")
    saas_readiness = data["metadata"].get("saas_readiness")
    if saas_readiness:
        lines.append(f"## {saas_readiness['title']}")
        lines.append("")
        lines.append("**Al aanwezig**")
        for bullet in saas_readiness.get("present", []):
            lines.append(f"- {bullet}")
        lines.append("")
        lines.append("**Deels aanwezig**")
        for bullet in saas_readiness.get("partial", []):
            lines.append(f"- {bullet}")
        lines.append("")
        lines.append("**Ontbreekt nog**")
        for bullet in saas_readiness.get("missing", []):
            lines.append(f"- {bullet}")
        lines.append("")
        lines.append("**Pas later doen**")
        for bullet in saas_readiness.get("later", []):
            lines.append(f"- {bullet}")
        lines.append("")

    post_checklist_system = data["metadata"].get("post_checklist_system")
    if post_checklist_system:
        lines.append(f"## {post_checklist_system['title']}")
        lines.append("")
        for bullet in post_checklist_system.get("bullets", []):
            lines.append(f"- {bullet}")
        lines.append("")

    title_by_prompt = {item["prompt_file"]: item["title"] for item in data["checklist"]["items"]}
    for phase in phases:
        lines.append(f"## {phase['title']}")
        lines.append("")
        lines.append("**Doel**")
        lines.append(phase["objective"])
        lines.append("")
        lines.append("**Hoort in deze fase**")
        for key in phase["items"]:
            title = title_by_prompt[key]
            row = rows_by_prompt.get(key)
            if row and row[2] == "Voldaan":
                lines.append(f"- {title} - afgerond")
            else:
                lines.append(f"- {title}")
        lines.append("")
        lines.append("**Status**")
        lines.extend(phase_status(rows_by_prompt, phase["items"], current_phase_id, phase["id"]))
        lines.append("")
        lines.append("**Exit gate**")
        for bullet in phase["exit_gate"]:
            lines.append(f"- {bullet}")
        if phase["waits_until_later"]:
            lines.append("")
            lines.append("**Wacht expliciet tot later**")
            for bullet in phase["waits_until_later"]:
                lines.append(f"- {bullet}")
        lines.append("")

    lines.append("## Externe afhankelijkheden en gates")
    lines.append("")
    for bullet in data["metadata"]["external_dependencies"]:
        lines.append(f"- {bullet}")
    lines.append("")
    lines.append("## Wat we nu bewust niet eerst doen")
    lines.append("")
    for bullet in data["metadata"]["not_now"]:
        lines.append(f"- {bullet}")
    lines.append("")
    lines.append("## Documentregels")
    lines.append("")
    lines.append('- gebruik de checklist voor "wat nu als eerst?"')
    lines.append('- gebruik deze roadmap voor "waarom in deze fase?"')
    lines.append("- gebruik externe documenten als input of referentie, niet als autonome werkvolgorde")
    lines.append("")
    return "\n".join(lines)


def verify_prompt_files(data: dict[str, Any]) -> list[str]:
    prompts_dir = ROOT / "docs" / "prompts"
    missing: list[str] = []
    for item in data["checklist"]["items"]:
        if not (prompts_dir / item["prompt_file"]).exists():
            missing.append(item["prompt_file"])
    return missing


def main() -> None:
    data = load_data()
    existing = load_existing_checklist_state()
    rows = build_checklist_rows(data, existing)
    save_checklist(data, rows)
    save_roadmap_workbook(data, rows)
    ROADMAP_PATH.write_text(build_roadmap_markdown(data, rows), encoding="utf-8")

    missing = verify_prompt_files(data)
    warnings = release_status_warnings(rows)
    print(f"Synced checklist: {CHECKLIST_PATH}")
    print(f"Synced roadmap:   {ROADMAP_PATH}")
    print(f"Synced workbook:  {ROADMAP_WORKBOOK_PATH}")
    if missing:
        print("Missing prompt files:")
        for item in missing:
            print(f"- {item}")
    else:
        print("All prompt files referenced in the checklist exist.")
    if warnings:
        print("Checklist release-status warnings:")
        for item in warnings:
            print(f"- {item}")


if __name__ == "__main__":
    main()
